import openpyxl
import re
from Common_Function import *

class ExcelHandler:

    def __init__(self, excel_dir:str):
        self.excel_dir = excel_dir
        self.wb_obj = openpyxl.load_workbook(self.excel_dir)

    def get_SWTS_DOORS(self, SWTS_DOORS_info:dict, TRS_start_string:str):
        test_case_info_dict = {}
        TITLE_ROW_NUM = SWTS_DOORS_info["sheet_title_row_num"]
        START_DATA_ROW_NUM = SWTS_DOORS_info["sheet_start_data_row_num"]
        title_col_num = {
                    "ID":0,
                    "Links from /SUZUKI/06_TEST/SW_TS_MSIL_YY8 to /SUZUKI/04_TRS/SWTRS_YY8_163D":0,
                    "Absolute Number":0,
                    "Relate To":0,
                    "Software Test Specification":0,
                    "TC_Content_Type":0,
                    "TC_General_Status":0,
                    "TC_Execution_Type":0,
                    "TC_Type":0,
                    "TC_Variant":0,
                    "TC_Precondition":0,
                    "TC_Steps":0,
                    "TC_Expected Results":0,
                    "TC Postcondition":0,
                    "TC_SW_Version":0,
                    "TC_IssueID":0,
                    "TC_Comment_Not_for_Customer":0,
                    "TC_Comment":0,
                    "HB_Functional_Safety":0
                    }

        for ws_obj in self.wb_obj.worksheets:
            if ws_obj.title == SWTS_DOORS_info["sheet_name"]:
                title_col_num = self._get_title_col_num(ws_obj=ws_obj, col_to_get=title_col_num, title_row_num=TITLE_ROW_NUM)
                row_num = START_DATA_ROW_NUM
                for row_num in range(START_DATA_ROW_NUM, ws_obj.max_row+1): 
                    TC_ID = ws_obj.cell(row=row_num,column=title_col_num["ID"]).value
                    if TC_ID not in test_case_info_dict.keys():
                        test_case_info_dict[TC_ID] = {}

                    Linked_TRS = ws_obj.cell(row=row_num,column=title_col_num["Links from /SUZUKI/06_TEST/SW_TS_MSIL_YY8 to /SUZUKI/04_TRS/SWTRS_YY8_163D"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Absolute Number"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Relate To"]).value
                    TC_name = ws_obj.cell(row=row_num,column=title_col_num["Software Test Specification"]).value
                    object_type = ws_obj.cell(row=row_num,column=title_col_num["TC_Content_Type"]).value
                    TC_status = ws_obj.cell(row=row_num,column=title_col_num["TC_General_Status"]).value
                    TC_execution_type = ws_obj.cell(row=row_num,column=title_col_num["TC_Execution_Type"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["TC_Type"]).value
                    TC_variant = ws_obj.cell(row=row_num,column=title_col_num["TC_Variant"]).value
                    TC_precondtion = ws_obj.cell(row=row_num,column=title_col_num["TC_Precondition"]).value
                    TC_execution_step = ws_obj.cell(row=row_num,column=title_col_num["TC_Steps"]).value
                    TC_expected_result = ws_obj.cell(row=row_num,column=title_col_num["TC_Expected Results"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["TC Postcondition"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["TC_SW_Version"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["TC_IssueID"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["TC_Comment_Not_for_Customer"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["TC_Comment"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Functional_Safety"]).value

                    test_case_info_dict[TC_ID]["Linked_TRS"] = self._get_linked_TRS(Linked_TRS, TRS_start_string)
                    test_case_info_dict[TC_ID]["TC_name"] = TC_name
                    test_case_info_dict[TC_ID]["object_type"] = object_type
                    test_case_info_dict[TC_ID]["TC_status"] = TC_status
                    test_case_info_dict[TC_ID]["TC_execution_type"] = TC_execution_type
                    test_case_info_dict[TC_ID]["TC_variant"] = TC_variant
                    test_case_info_dict[TC_ID]["TC_precondtion"] = TC_precondtion
                    test_case_info_dict[TC_ID]["TC_execution_step"] = TC_execution_step
                    test_case_info_dict[TC_ID]["TC_expected_result"] = TC_expected_result

                    row_num += 1

        return test_case_info_dict

    def get_TRS_DOORS(self, TRS_DOORS_info:dict):
        TRS_info_dict = {}
        TITLE_ROW_NUM = TRS_DOORS_info["sheet_title_row_num"]
        START_DATA_ROW_NUM = TRS_DOORS_info["sheet_start_data_row_num"]
        title_col_num = {
                    "ID":0,
                    "Last Modified On":0,
                    "Links to /SUZUKI/04_TRS/SWTRS_YY8_163D from /SUZUKI/06_TEST/SW_TS_MSIL_YY8":0,
                    "HB_FeatureNumber":0,
                    "MCU and DSP SWTRS for YY8":0,
                    "HB_VerificationCriteria":0,
                    "HB_VerificationMethod":0,
                    "HB_Project_Variant":0,
                    "HB_Responsible_Domain":0,
                    "HB_ContentType":0,
                    "HB_General_Status":0,
                    "HB_Milestone":0,
                    "HB_Implementation_Status":0,
                    "HB_Comment":0,
                    "HB_Review_Ticket_ID":0
                    }

        for ws_obj in self.wb_obj.worksheets:
            if ws_obj.title == TRS_DOORS_info["sheet_name"]:
                title_col_num = self._get_title_col_num(ws_obj=ws_obj, col_to_get=title_col_num, title_row_num=TITLE_ROW_NUM)
                row_num = START_DATA_ROW_NUM
                for row_num in range(START_DATA_ROW_NUM, ws_obj.max_row+1): 
                    TRS_ID = ws_obj.cell(row=row_num,column=title_col_num["ID"]).value
                    if TRS_ID not in TRS_info_dict.keys():
                        TRS_info_dict[TRS_ID] = {}

                    # ws_obj.cell(row=row_num,column=title_col_num["Last Modified On"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Links to /SUZUKI/04_TRS/SWTRS_YY8_163D from /SUZUKI/06_TEST/SW_TS_MSIL_YY8"]).value
                    feature_num = ws_obj.cell(row=row_num,column=title_col_num["HB_FeatureNumber"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["MCU and DSP SWTRS for YY8"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_VerificationCriteria"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_VerificationMethod"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Project_Variant"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Responsible_Domain"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_ContentType"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_General_Status"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Milestone"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Implementation_Status"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Comment"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["HB_Review_Ticket_ID"]).value

                    TRS_info_dict[TRS_ID]["feature_num"] = feature_num

                    row_num += 1

        return TRS_info_dict

    def read_req_coverage_template_report(self, template_report_info:dict):
        req_coverage_info_dict = {}
        TITLE_ROW_NUM = template_report_info["sheet_requirement_coverage_title_row_num"]
        START_DATA_ROW_NUM = template_report_info["sheet_requirement_coverage_start_data_row_num"]
        title_col_num = {
                        "FROP":0,
                        "Feature":0,
                        "Required Feature":0,
                        "Total Test Case":0,
                        "Pass":0,
                        "Fail":0,
                        "Not Tested":0,
                        "Pass percentage":0,
                        "Total Requirements":0,
                        "Tested Requirements":0,
                        "Mile Stone SW Test Coverage":0,
                        "SW Test Coverage Reference":0
                        }

        for ws_obj in self.wb_obj.worksheets:
            if ws_obj.title == req_coverage_info_dict["sheet_name_requirement_coverage"]:
                title_col_num = self._get_title_col_num(ws_obj=ws_obj, col_to_get=title_col_num, title_row_num=TITLE_ROW_NUM)
                row_num = START_DATA_ROW_NUM
                for row_num in range(START_DATA_ROW_NUM, ws_obj.max_row+1): 
                    FROP_ID = ws_obj.cell(row=row_num,column=title_col_num["FROP"]).value
                    if FROP_ID not in req_coverage_info_dict.keys():
                        req_coverage_info_dict[FROP_ID] = {}

                    feature_name = ws_obj.cell(row=row_num,column=title_col_num["Feature"]).value
                    required_feature = ws_obj.cell(row=row_num,column=title_col_num["Required Feature"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Total Test Case"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Pass"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Fail"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Not Tested"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Pass percentage"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Total Requirements"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Tested Requirements"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Mile Stone SW Test Coverage"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["SW Test Coverage Reference"]).value


                    req_coverage_info_dict[FROP_ID]["feature_name"] = feature_name
                    req_coverage_info_dict[FROP_ID]["required_feature"] = self._get_required_feature(required_feature)

                    row_num += 1

        return req_coverage_info_dict
    
    def fill_test_report_save_to_output(self, template_report_info:dict, database_handler, output_dir:str):
        test_report_info_dict = {}
        TITLE_ROW_NUM = template_report_info["sheet_test_report_title_row_num"]
        START_DATA_ROW_NUM = template_report_info["sheet_test_report_start_data_row_num"]
        title_col_num = {
                        "ID":0,
                        "Feature name":0,
                        "Sub-feature":0,
                        "Author":0,
                        "Requirement Source":0,
                        "Automation":0,
                        "Summary":0,
                        "Pre-condition":0,
                        "Test Procedure":0,
                        "Expected results":0,
                        "Observed results":0,
                        "Result":0,
                        "Executor":0,
                        "Date":0,
                        "Remarks":0,
                        "YY8":0,
                        "Y163D":0
                        }       
        
        SWTS_database = database_handler.SWTS_DOORS_database
        TRS_database = database_handler.TRS_DOORS_database
        req_coverage_database = database_handler.req_coverage_database

        for ws_obj in self.wb_obj.worksheets:
            if ws_obj.title == test_report_info_dict["sheet_name_test_report"]:
                title_col_num = self._get_title_col_num(ws_obj=ws_obj, col_to_get=title_col_num, title_row_num=TITLE_ROW_NUM)
                row_num = START_DATA_ROW_NUM
                for TC_ID in SWTS_database.keys(): 
                    if SWTS_database[TC_ID]["object_type"] != "Test Case":
                        continue
                    
                    ws_obj.cell(row=row_num,column=title_col_num["ID"]).value = TC_ID
                    ws_obj.cell(row=row_num,column=title_col_num["Feature name"]).value = self._get_TC_feature_name(SWTS_database[TC_ID]["Linked_TRS"],
                                                                                                                    TRS_database,
                                                                                                                    req_coverage_database)
                    # ws_obj.cell(row=row_num,column=title_col_num["Sub-feature"]).value
                    ws_obj.cell(row=row_num,column=title_col_num["Author"]).value = "Thanh Tran"
                    ws_obj.cell(row=row_num,column=title_col_num["Requirement Source"]).value = ListStringToStringEndLine(SWTS_database[TC_ID]["Linked_TRS"])
                    ws_obj.cell(row=row_num,column=title_col_num["Automation"]).value = self._get_automation_stage(SWTS_database[TC_ID]["TC_execution_type"])
                    ws_obj.cell(row=row_num,column=title_col_num["Summary"]).value = SWTS_database[TC_ID]["TC_name"]
                    ws_obj.cell(row=row_num,column=title_col_num["Pre-condition"]).value = SWTS_database[TC_ID]["TC_precondtion"]
                    ws_obj.cell(row=row_num,column=title_col_num["Test Procedure"]).value = SWTS_database[TC_ID]["TC_execution_step"]
                    ws_obj.cell(row=row_num,column=title_col_num["Expected results"]).value = SWTS_database[TC_ID]["TC_expected_result"]
                    # ws_obj.cell(row=row_num,column=title_col_num["Observed results"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Result"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Executor"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Date"]).value
                    # ws_obj.cell(row=row_num,column=title_col_num["Remarks"]).value
                    ws_obj.cell(row=row_num,column=title_col_num["YY8"]).value = self._get_support_variant("YY8", SWTS_database[TC_ID]["TC_variant"])
                    ws_obj.cell(row=row_num,column=title_col_num["Y163D"]).value = self._get_support_variant("Y163D", SWTS_database[TC_ID]["TC_variant"])

                    row_num += 1

        self.wb_obj.save(f"{output_dir}\\MSIL_YY8_SW_QualificationTestReport_VERSION.xlsx")
        self.wb_obj.save(f"{output_dir}\\MSIL_163D_SW_QualificationTestReport_VERSION.xlsx")

    def _get_TC_feature_name(self, TC_TRS_linked_list:list, TRS_database:dict, req_coverage_database:dict):
        feature_num_list = []
        TC_feature_name_output = ""
        
        for TRS in TC_TRS_linked_list:
            if TRS_database[TRS]["feature_num"] not in feature_num_list:
                feature_num_list.append(TRS_database[TRS]["feature_num"])

        for feature_num in feature_num_list:
            if (req_coverage_database[feature_num]["required_feature"] == True
            and req_coverage_database[feature_num]["feature_name"] not in TC_feature_name_output):
                TC_feature_name_output += req_coverage_database[feature_num]["feature_name"] + "\n"

        return TC_feature_name_output
    
    def _get_support_variant(self, current_variant:str, TC_variant:str):
        if TC_variant.lower() == "general":
            return "Yes"
        # Need to update when have test cases for specific variant
        else:
            return "No"
    
    def _get_automation_stage(self, input_string_stage:str):
        if input_string_stage.lower() == "manual":
            return "No"
        else:
            return "Yes"
    
    def _get_required_feature(self, input_string:str):
        if input_string.lower() == "yes":
            return True
        else:
            return False
    
    def _get_linked_TRS(self, input_link_string:str, TRS_start_string:str):
        output_list = []
        regex_find_TRS_num = f"\({TRS_start_string}\d+\)"

        output_list = re.findall(regex_find_TRS_num, input_link_string)

        return output_list

    def _get_title_col_num(self, ws_obj, col_to_get:dict, title_row_num):
        """
        @Description: get all the col num for all the title in excel sheet
        """
        for col_num in range(1, ws_obj.max_column+1):
            if ws_obj.cell(row=title_row_num,column=col_num).value in col_to_get.keys():
                    col_to_get[ws_obj.cell(row=1,column=col_num).value] = col_num
        for key in col_to_get.keys():
            if col_to_get[key] == 0:
                print(f"column '{key}' is missing in the input excel report!")
        return col_to_get              
            
